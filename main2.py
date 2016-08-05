# -*- coding: utf-8 -*-
import sys
import getpath
import os
import re
from openpyxl import load_workbook
def mylistdir(p,f):
    a=os.listdir(p)
    fs=myfind(a,f)
    return(fs)
def myfind(l,p):
    lr=[];
    #print p
    p1=p.replace(".",r"\.")
    p2=p1.replace("*",".*")
    p2=p2+"$"
    p2="^"+p2
    for a in l:
        #print a
        if  re.search(p2,a,re.IGNORECASE)==None :
           pass
           #print "pass"
        else:
           lr.append(a)
       #print "append"
    return lr
def main2(fn,fn2):
	xlBook = load_workbook(filename = fn)
	data={}
	table=xlBook.worksheets[0]  
	rows=len(table.rows)
	for i in range(rows)[1:]:
		data[table.cell(row=i+1,column=1).value]=table.cell(row=i+1,column=5).value
	table=xlBook.worksheets[1]  
	rows=len(table.rows)
	for i in range(rows)[1:]:
		data[table.cell(row=i+1,column=1).value]=table.cell(row=i+1,column=5).value
	wb = load_workbook(filename = fn2)
	sheet=wb.get_sheet_by_name("Sheet1")
	rows=len(sheet.rows)
	for i in range(rows)[4:]:
		sheet.cell(row = i+1, column = 18).value=data.get(sheet.cell(row = i+1, column = 4).value)
	wb.save(filename = fn2)
def main():
    initpath=os.path.abspath(".")+"\\"
    fs=mylistdir(initpath,"现存量*.xlsx")
    if len(fs)>0:
        fn=fs[0]
    else:
        print("未发现 现存量*.xlsx")    
        return 
    fs=mylistdir(initpath,"采购Bom*.xlsx")
    if len(fs)>0:
        fn2=fs[0]
    else:
        print("未发现 采购Bom*.xlsx")    
        return
    main2(initpath+fn,initpath+fn2)
if __name__=="__main__":
    main()
